"""
Module 9 — Catalogue multi-format (PDF, HTML, Excel, JSON, XML)
Pays = Algérie STATIQUE sur tous les fournisseurs.
Accepte : suppliers liste OU dict unique (depuis M4/M5).

Lancer seul :  python modules/module_9_catalogue.py
"""

import os
import json
import datetime
import xml.etree.ElementTree as ET
from xml.dom import minidom

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                 Table, TableStyle, HRFlowable)

try:
    from jinja2 import Environment
    HAS_JINJA = True
except ImportError:
    HAS_JINJA = False

# ── Constante pays ─────────────────────────────────────────────────────────
PAYS = "Algérie"  # STATIQUE

# ── Données mock ────────────────────────────────────────────────────────────
MOCK_SPECS = {
    "diametre_nominal": "DN100",
    "pression_nominale": "PN40",
    "materiau": {
        "corps": "Fonte GS 400-15",
        "disque": "Inox 316L",
        "siege": "PTFE",
        "tige": "Inox 316",
        "joint": "EPDM"
    },
    "longueur_face_a_face": "229 mm",
    "tolerance": "±0.1 mm",
    "norme": "EN 558",
    "plan_dxf": "généré ✓"
}

MOCK_SUPPLIERS = [
    {"nom_fournisseur": "ValveTech SRL",   "pays": PAYS, "prix_unitaire": 4200,
     "email": "sales@valvetech.dz",      "delai_livraison": "18 jours"},
    {"nom_fournisseur": "HydroVal Algérie", "pays": PAYS, "prix_unitaire": 5500,
     "email": "hydro@hydroval.dz",       "delai_livraison": "12 jours"},
    {"nom_fournisseur": "AlgéroFond SARL", "pays": PAYS, "prix_unitaire": 4800,
     "email": "contact@algerofond.dz",   "delai_livraison": "14 jours"},
]

MOCK_SWOT = {
    "forces": "Matériau premium 316L, normes EN respectées, réseau algérien établi.",
    "faiblesses": "Coût unitaire élevé, délai livraison variable.",
    "opportunites": "Croissance industrie pétro-gaz Algérie, substitution importations.",
    "menaces": "Fluctuation DZD, concurrence asiatique, réglementation douanière.",
}

MOCK_FINANCE = {
    "roi_pct": 34,
    "van_dzd": 280000,
    "tri_pct": 28,
    "tco_total": 847250,
    "inflation_pct": 5.2,
    "maintenance_annuelle": 42000,
}


# ── Utilitaires ──────────────────────────────────────────────────────────
def _normalize_suppliers(raw) -> list:
    """
    Accepte liste ou dict unique, force pays = Algérie sur tous.
    Normalise aussi 'delai' → 'delai_livraison'.
    """
    if isinstance(raw, dict):
        raw = [raw]
    if not raw:
        return MOCK_SUPPLIERS
    result = []
    for s in raw:
        s = dict(s)
        s["pays"] = PAYS  # STATIQUE
        # Normaliser la clé délai
        if "delai" in s and "delai_livraison" not in s:
            s["delai_livraison"] = s.pop("delai")
        result.append(s)
    return result


def _flatten_specs(specs: dict) -> list:
    items = []
    for k, v in specs.items():
        if isinstance(v, dict):
            for sk, sv in v.items():
                items.append((f"{k} — {sk}", str(sv)))
        else:
            items.append((k.replace("_", " ").capitalize(), str(v)))
    return items


def _best_supplier(suppliers: list) -> str:
    if not suppliers:
        return ""
    return min(suppliers, key=lambda s: float(s.get("prix_unitaire", 9e9))
               ).get("nom_fournisseur", "")


def _tco_bars(finance: dict, n_years: int = 5) -> list:
    base = float(finance.get("maintenance_annuelle", 42000))
    infl = float(finance.get("inflation_pct", 5.2)) / 100
    year = datetime.datetime.now().year
    return [{"annee": str(year + i),
             "valeur": round(base * ((1 + infl) ** i))}
            for i in range(n_years)]


def _get_delai(s: dict) -> str:
    return s.get("delai_livraison") or s.get("delai") or "N/A"


# ── 1. PDF ───────────────────────────────────────────────────────────────
def generate_catalogue_pdf(specs, suppliers, finance, swot, output_path):
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            rightMargin=20*mm, leftMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)
    sty = getSampleStyleSheet()
    W = 170 * mm

    cover_badge = ParagraphStyle("CvBg",   parent=sty["Normal"], fontSize=9,
                                  textColor=colors.HexColor("#7eb8e8"), spaceAfter=8)
    cover_title = ParagraphStyle("CvTi",   parent=sty["Normal"], fontSize=22,
                                  textColor=colors.white, fontName="Helvetica-Bold", spaceAfter=4)
    cover_sub   = ParagraphStyle("CvSb",   parent=sty["Normal"], fontSize=12,
                                  textColor=colors.HexColor("#93aec6"), spaceAfter=12)
    section_hd  = ParagraphStyle("ScHd",   parent=sty["Normal"], fontSize=11,
                                  textColor=colors.HexColor("#2c3e50"),
                                  fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)
    swot_lbl    = ParagraphStyle("SwLb",   parent=sty["Normal"], fontSize=9,
                                  fontName="Helvetica-Bold", spaceAfter=2)
    normal      = ParagraphStyle("N",      parent=sty["Normal"], fontSize=9, spaceAfter=3)
    footer_s    = ParagraphStyle("Ft",     parent=sty["Normal"], fontSize=8,
                                  textColor=colors.HexColor("#7f8c8d"), alignment=1)

    DARK_COVER = colors.HexColor("#1a2a3a")
    DARK_HDR   = colors.HexColor("#34495e")
    LIGHT_ROW  = colors.HexColor("#f8f9fa")
    BADGE_GRN  = colors.HexColor("#d4edda")
    TXT_GRN    = colors.HexColor("#155724")
    KPI_BG     = colors.HexColor("#f0f4f8")
    BAR_BLUE   = colors.HexColor("#378add")

    story = []
    dn = specs.get("diametre_nominal", "DN100")
    pn = specs.get("pression_nominale", "PN40")
    mat = specs.get("materiau", {})
    mat_str = mat.get("disque", "Inox 316L") if isinstance(mat, dict) else str(mat)
    best = _best_supplier(suppliers)

    # ── Couverture ─────────────────────────────────────────────────────
    cv = Table([
        [Paragraph(f"OpenIndustry {PAYS} · Module 9 — Catalogue", cover_badge)],
        [Paragraph(f"Vanne industrielle {dn} {pn}", cover_title)],
        [Paragraph(f"{mat_str} · Réf. VAN-{dn}-{pn} · {specs.get('norme','EN 558')} · "
                   f"{datetime.datetime.now().year}", cover_sub)],
    ], colWidths=[W])
    cv.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), DARK_COVER),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 2), (-1, 2), 18),
        ("LEFTPADDING", (0, 0), (-1, -1), 16),
        ("RIGHTPADDING", (0, 0), (-1, -1), 16),
    ]))
    story.extend([cv, Spacer(1, 8*mm)])

    # ── Fiche technique ────────────────────────────────────────────────
    story.append(Paragraph("Fiche technique", section_hd))
    flat = _flatten_specs(specs)
    spec_data = [[Paragraph(f"<b>{k}</b>", normal), Paragraph(v, normal)]
                 for k, v in flat]
    spec_table = Table(spec_data, colWidths=[70*mm, W-70*mm])
    spec_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bdc3c7")),
        ("BACKGROUND", (0, 0), (0, -1), KPI_BG),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, LIGHT_ROW]),
    ]))
    story.extend([spec_table, Spacer(1, 6*mm)])

    # ── Fournisseurs ───────────────────────────────────────────────────
    story.append(Paragraph("Fournisseurs qualifiés", section_hd))
    sup_hdr = [Paragraph(f"<b>{h}</b>", normal)
               for h in ["Fournisseur", "Pays", "Email", "Prix unit. (DZD)",
                         "Délai", "TCO total (DZD)"]]
    sup_data = [sup_hdr]
    for s in suppliers[:6]:
        nom = s.get("nom_fournisseur", "")
        badge = " ★ Meilleur" if nom == best else ""
        sup_data.append([
            Paragraph(f"<b>{nom}</b>{badge}", normal),
            Paragraph(PAYS, normal),  # STATIQUE
            Paragraph(s.get("email", ""), normal),
            Paragraph(f"{float(s.get('prix_unitaire', 0)):,.0f}", normal),
            Paragraph(_get_delai(s), normal),
            Paragraph(f"{float(s.get('tco_total', 0)):,.0f}", normal),
        ])
    col_w = [38*mm, 22*mm, 40*mm, 24*mm, 20*mm, W-144*mm]
    sup_table = Table(sup_data, colWidths=col_w, repeatRows=1)
    sup_style = [
        ("BACKGROUND", (0, 0), (-1, 0), DARK_HDR),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bdc3c7")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_ROW]),
    ]
    for i, s in enumerate(suppliers[:6], 1):
        if s.get("nom_fournisseur") == best:
            sup_style += [("BACKGROUND", (0, i), (-1, i), BADGE_GRN),
                           ("TEXTCOLOR", (0, i), (-1, i), TXT_GRN)]
    sup_table.setStyle(TableStyle(sup_style))
    story.extend([sup_table, Spacer(1, 6*mm)])

    # ── TCO KPIs ───────────────────────────────────────────────────────
    story.append(Paragraph("TCO & Indicateurs financiers", section_hd))
    kpi_data = [
        [Paragraph(f"<b>{float(finance.get('tco_total', 0)):,.0f} DZD</b>", normal),
         Paragraph(f"<b>{finance.get('inflation_pct', 0)} %</b>", normal),
         Paragraph(f"<b>{float(finance.get('maintenance_annuelle', 0)):,.0f} DZD</b>", normal)],
        [Paragraph("TCO total", normal),
         Paragraph("Inflation moy. (World Bank)", normal),
         Paragraph("Maintenance / an", normal)],
    ]
    kpi_table = Table(kpi_data, colWidths=[W/3]*3)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), KPI_BG),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d0d8e0")),
    ]))
    story.extend([kpi_table, Spacer(1, 4*mm)])

    # Barres TCO
    bars = _tco_bars(finance)
    max_val = max(r["valeur"] for r in bars)
    bar_data = []
    for r in bars:
        pct = r["valeur"] / max_val
        bar_w = pct * (W - 80*mm)
        bar_row = [
            Paragraph(r["annee"], normal),
            Table([[""]], colWidths=[max(bar_w, 5*mm)], rowHeights=[8*mm],
                  style=[("BACKGROUND", (0, 0), (-1, -1), BAR_BLUE)]),
            Paragraph(f"{r['valeur']:,.0f} DZD", normal),
        ]
        bar_data.append(bar_row)
    bar_table = Table(bar_data, colWidths=[18*mm, W-80*mm, 60*mm])
    bar_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.extend([bar_table, Spacer(1, 6*mm)])

    # ── SWOT ──────────────────────────────────────────────────────────
    story.append(Paragraph("Analyse SWOT", section_hd))
    swot_items = [
        ("Forces",       str(swot.get("forces", "")),       "#eaf3de", "#3b6d11"),
        ("Faiblesses",   str(swot.get("faiblesses", "")),   "#faece7", "#993c1d"),
        ("Opportunités", str(swot.get("opportunites", "")), "#e6f1fb", "#185fa5"),
        ("Menaces",      str(swot.get("menaces", "")),      "#faeeda", "#854f0b"),
    ]
    half = W / 2 - 2*mm
    swot_rows = []
    for i in range(0, 4, 2):
        row = []
        for j in range(2):
            lbl, txt, bg, tc = swot_items[i+j]
            if isinstance(txt, list):
                txt = " · ".join(txt)
            cell = Table([
                [Paragraph(f"<b>{lbl}</b>",
                           ParagraphStyle("sl", parent=swot_lbl,
                                          textColor=colors.HexColor(tc)))],
                [Paragraph(txt, ParagraphStyle("st", parent=normal, spaceAfter=0))],
            ], colWidths=[half])
            cell.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(bg)),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ]))
            row.append(cell)
        swot_rows.append(row)
    swot_table = Table(swot_rows, colWidths=[half, half])
    swot_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.extend([swot_table, Spacer(1, 6*mm)])

    # ── KPIs financiers ────────────────────────────────────────────────
    story.append(Paragraph("Indicateurs financiers", section_hd))
    fin_data = [
        [Paragraph(f"<b>{finance.get('roi_pct', 0)} %</b>", normal),
         Paragraph(f"<b>+{float(finance.get('van_dzd', 0)):,.0f}</b>", normal),
         Paragraph(f"<b>{finance.get('tri_pct', 0)} %</b>", normal)],
        [Paragraph("ROI (3 ans)", normal),
         Paragraph("VAN (DZD)", normal),
         Paragraph("TRI", normal)],
    ]
    fin_table = Table(fin_data, colWidths=[W/3]*3)
    fin_ts = [
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d0d8e0")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#d4edda")),
        ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#d1ecf1")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f0f4f8")),
    ]
    fin_table.setStyle(TableStyle(fin_ts))
    story.extend([fin_table, Spacer(1, 8*mm)])

    # ── Pied de page ───────────────────────────────────────────────────
    story.append(HRFlowable(width=W, color=colors.HexColor("#bdc3c7"), thickness=0.5))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        f"Document généré automatiquement par INDUSTRIE IA · "
        f"OpenIndustry {PAYS} · {datetime.datetime.now().strftime('%d/%m/%Y')}",
        footer_s))

    doc.build(story)
    print(f"[M9] ✓ PDF : {output_path}")
    return output_path


# ── 2. HTML (Jinja2 inline) ───────────────────────────────────────────────
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<title>Catalogue — {{ specs.get('diametre_nominal','') }} {{ specs.get('pression_nominale','') }}</title>
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#f5f7fa;color:#2c3e50;padding:2rem}
.page{background:#fff;border:1px solid #dde3ea;border-radius:12px;overflow:hidden;max-width:820px;margin:0 auto}
.cover{background:#1a2a3a;color:#fff;padding:32px 36px 24px}
.cover-badge{font-size:11px;letter-spacing:.08em;text-transform:uppercase;color:#7eb8e8;margin-bottom:10px}
.cover-title{font-size:24px;font-weight:600;line-height:1.2;margin-bottom:4px}
.cover-sub{font-size:13px;color:#93aec6;margin-bottom:18px}
.cover-tags{display:flex;flex-wrap:wrap;gap:8px}
.cover-tag{font-size:11px;padding:4px 12px;border-radius:20px;background:rgba(255,255,255,.1);color:#c8ddf0}
.section{padding:20px 28px;border-bottom:1px solid #eaecee}
.section:last-child{border-bottom:none}
.section-title{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.07em;color:#7f8c8d;margin-bottom:12px}
.specs-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px}
.spec-item{background:#f0f4f8;border-radius:8px;padding:10px 14px}
.spec-label{font-size:11px;color:#95a5a6;margin-bottom:3px}
.spec-value{font-size:14px;font-weight:600}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#34495e;color:#fff;padding:8px 10px;text-align:left;font-weight:600}
td{padding:8px 10px;border-bottom:1px solid #eaecee}
tr:last-child td{border-bottom:none}
tr.best td{background:#d4edda;color:#155724}
.badge{display:inline-block;font-size:10px;padding:2px 8px;border-radius:20px;background:#155724;color:#fff;margin-left:6px;vertical-align:middle}
.tco-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:14px}
.tco-card{background:#f0f4f8;border-radius:8px;padding:12px 14px;text-align:center}
.tco-val{font-size:18px;font-weight:600}
.tco-lbl{font-size:11px;color:#7f8c8d;margin-top:3px}
.bar-row{display:flex;align-items:center;gap:10px;margin-bottom:6px;font-size:12px}
.bar-year{min-width:38px;color:#7f8c8d}
.bar-bg{flex:1;height:10px;background:#eaecee;border-radius:6px;overflow:hidden}
.bar-fill{height:100%;border-radius:6px;background:#378add}
.bar-val{min-width:110px;text-align:right;color:#7f8c8d}
.swot-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px}
.swot-card{border-radius:8px;padding:12px 14px}
.swot-card.s{background:#eaf3de}.swot-card.s .swot-lbl{color:#3b6d11}
.swot-card.w{background:#faece7}.swot-card.w .swot-lbl{color:#993c1d}
.swot-card.o{background:#e6f1fb}.swot-card.o .swot-lbl{color:#185fa5}
.swot-card.t{background:#faeeda}.swot-card.t .swot-lbl{color:#854f0b}
.swot-lbl{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px}
.swot-text{font-size:12px;line-height:1.5;color:#2c3e50}
.fin-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:8px}
.fin-card{border-radius:8px;padding:12px 14px;text-align:center}
.fin-card.roi{background:#d4edda}.fin-card.roi .fin-val{color:#155724}
.fin-card.van{background:#d1ecf1}.fin-card.van .fin-val{color:#0c5460}
.fin-card.tri{background:#f0f4f8}.fin-card.tri .fin-val{color:#2c3e50}
.fin-val{font-size:20px;font-weight:700}
.fin-lbl{font-size:11px;color:#7f8c8d;margin-top:3px}
.fmt-row{display:flex;flex-wrap:wrap;gap:8px}
.fmt-pill{font-size:12px;font-weight:600;padding:6px 16px;border-radius:20px;border:1px solid #dde3ea}
.fmt-pdf{background:#fce8e8;color:#922b21}.fmt-html{background:#fde8d0;color:#784212}
.fmt-xlsx{background:#d4edda;color:#155724}.fmt-json{background:#d1ecf1;color:#0c5460}
.fmt-xml{background:#e8d8fa;color:#512e5f}
footer{padding:14px 28px;font-size:11px;color:#95a5a6;text-align:center;border-top:1px solid #eaecee}
</style>
</head>
<body>
<div class="page">
  <div class="cover">
    <div class="cover-badge">OpenIndustry {{ PAYS }} · Module 9 — Catalogue</div>
    <div class="cover-title">Vanne industrielle {{ specs.get('diametre_nominal','') }} {{ specs.get('pression_nominale','') }}</div>
    <div class="cover-sub">{{ flat_material }} · Réf. VAN-{{ specs.get('diametre_nominal','') }}-{{ specs.get('pression_nominale','') }} · {{ specs.get('norme','EN 558') }} · {{ year }}</div>
    <div class="cover-tags">
      <span class="cover-tag">{{ PAYS }}</span>
      <span class="cover-tag">{{ nb_suppliers }} fournisseurs</span>
      <span class="cover-tag">Livraison {{ min_delai }}</span>
      <span class="cover-tag">TCO 10 ans calculé</span>
    </div>
  </div>
  <div class="section">
    <div class="section-title">Fiche technique</div>
    <div class="specs-grid">
      {% for lbl, val in flat_specs %}
      <div class="spec-item"><div class="spec-label">{{ lbl }}</div><div class="spec-value">{{ val }}</div></div>
      {% endfor %}
    </div>
  </div>
  <div class="section">
    <div class="section-title">Fournisseurs qualifiés</div>
    <table>
      <thead><tr><th>Fournisseur</th><th>Pays</th><th>Email</th><th>Prix unit. (DZD)</th><th>Délai</th><th>TCO total (DZD)</th></tr></thead>
      <tbody>
        {% for s in suppliers %}
        <tr {% if s.nom_fournisseur == best_supplier %}class="best"{% endif %}>
          <td>{{ s.nom_fournisseur }}{% if s.nom_fournisseur == best_supplier %}<span class="badge">Meilleur</span>{% endif %}</td>
          <td><b>{{ PAYS }}</b></td>
          <td>{{ s.email }}</td>
          <td>{{ "{:,.0f}".format(s.prix_unitaire|float) }}</td>
          <td>{{ s.delai_livraison }}</td>
          <td>{{ "{:,.0f}".format(s.tco_total|float) }}</td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  <div class="section">
    <div class="section-title">TCO & Maintenance</div>
    <div class="tco-grid">
      <div class="tco-card"><div class="tco-val">{{ "{:,.0f}".format(finance.tco_total|float) }}</div><div class="tco-lbl">TCO total (DZD)</div></div>
      <div class="tco-card"><div class="tco-val">{{ finance.inflation_pct }} %</div><div class="tco-lbl">Inflation moy. (World Bank)</div></div>
      <div class="tco-card"><div class="tco-val">{{ "{:,.0f}".format(finance.maintenance_annuelle|float) }}</div><div class="tco-lbl">Maintenance / an (DZD)</div></div>
    </div>
    {% set max_bar = tco_bars | map(attribute='valeur') | max %}
    {% for row in tco_bars %}
    <div class="bar-row">
      <span class="bar-year">{{ row.annee }}</span>
      <div class="bar-bg"><div class="bar-fill" style="width:{{ (row.valeur / max_bar * 100) | round | int }}%"></div></div>
      <span class="bar-val">{{ "{:,.0f}".format(row.valeur) }} DZD</span>
    </div>
    {% endfor %}
  </div>
  <div class="section">
    <div class="section-title">Analyse SWOT</div>
    <div class="swot-grid">
      <div class="swot-card s"><div class="swot-lbl">Forces</div><div class="swot-text">{{ swot.forces }}</div></div>
      <div class="swot-card w"><div class="swot-lbl">Faiblesses</div><div class="swot-text">{{ swot.faiblesses }}</div></div>
      <div class="swot-card o"><div class="swot-lbl">Opportunités</div><div class="swot-text">{{ swot.opportunites }}</div></div>
      <div class="swot-card t"><div class="swot-lbl">Menaces</div><div class="swot-text">{{ swot.menaces }}</div></div>
    </div>
  </div>
  <div class="section">
    <div class="section-title">Indicateurs financiers</div>
    <div class="fin-grid">
      <div class="fin-card roi"><div class="fin-val">{{ finance.roi_pct }} %</div><div class="fin-lbl">ROI (3 ans)</div></div>
      <div class="fin-card van"><div class="fin-val">+{{ "{:,.0f}".format(finance.van_dzd|float) }}</div><div class="fin-lbl">VAN (DZD)</div></div>
      <div class="fin-card tri"><div class="fin-val">{{ finance.tri_pct }} %</div><div class="fin-lbl">TRI</div></div>
    </div>
  </div>
  <div class="section">
    <div class="section-title">5 formats de sortie — Module 9</div>
    <div class="fmt-row">
      <span class="fmt-pill fmt-pdf">PDF</span><span class="fmt-pill fmt-html">HTML</span>
      <span class="fmt-pill fmt-xlsx">Excel</span><span class="fmt-pill fmt-json">JSON</span>
      <span class="fmt-pill fmt-xml">XML</span>
    </div>
  </div>
  <footer>Généré par INDUSTRIE IA · OpenIndustry {{ PAYS }} · {{ date }}</footer>
</div>
</body>
</html>"""


def generate_catalogue_html(specs, suppliers, finance, swot, output_path):
    if not HAS_JINJA:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(f"<html><body><h1>Catalogue {specs.get('diametre_nominal')}</h1>"
                    "<p>Jinja2 non installé</p></body></html>")
        return output_path

    flat = _flatten_specs(specs)
    mat = specs.get("materiau", {})
    flat_material = mat.get("disque", "Inox 316L") if isinstance(mat, dict) else str(mat)
    best = _best_supplier(suppliers)
    bars = _tco_bars(finance)
    delais = [_get_delai(s) for s in suppliers]
    min_delai = sorted(delais, key=lambda d: int("".join(filter(str.isdigit, d)) or "99"))[0] if delais else "N/A"

    class AttrDict(dict):
        def __getattr__(self, k):
            v = self.get(k, "")
            return v

    # Normaliser swot pour Jinja (peut être list ou str)
    swot_norm = {}
    for k, v in swot.items():
        swot_norm[k] = " · ".join(v) if isinstance(v, list) else str(v)

    # Ajouter delai_livraison aux suppliers pour Jinja
    sup_jinja = []
    for s in suppliers:
        sd = dict(s)
        sd["delai_livraison"] = _get_delai(s)
        sd["pays"] = PAYS  # STATIQUE
        sup_jinja.append(sd)

    env = Environment()
    tpl = env.from_string(HTML_TEMPLATE)
    html = tpl.render(
        PAYS=PAYS,
        specs=specs,
        flat_specs=flat,
        flat_material=flat_material,
        suppliers=sup_jinja,
        best_supplier=best,
        finance=AttrDict(finance),
        swot=AttrDict(swot_norm),
        tco_bars=bars,
        nb_suppliers=len(suppliers),
        min_delai=min_delai,
        year=datetime.datetime.now().year,
        date=datetime.datetime.now().strftime("%d/%m/%Y"),
    )
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[M9] ✓ HTML : {output_path}")
    return output_path


# ── 3. Excel ─────────────────────────────────────────────────────────────
def generate_catalogue_excel(specs, suppliers, finance, swot, output_path):
    def _border():
        t = Side(border_style="thin", color="BDC3C7")
        return Border(left=t, right=t, top=t, bottom=t)

    def _fill(h):
        return PatternFill("solid", fgColor=h)

    wb = openpyxl.Workbook()

    # Onglet Catalogue
    ws = wb.active
    ws.title = "Catalogue"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 26

    ws.merge_cells("A1:B1")
    ws["A1"] = f"Catalogue Produit — OpenIndustry {PAYS}"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    ws["A1"].fill = _fill("1A2A3A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for r, txt in [(2, f"Référence : {specs.get('diametre_nominal','')} {specs.get('pression_nominale','')}"),
                   (3, f"Pays : {PAYS} · Date : {datetime.datetime.now().strftime('%d/%m/%Y')}")]:
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"] = txt
        ws[f"A{r}"].font = Font(color="93AEC6", name="Calibri", size=10)
        ws[f"A{r}"].fill = _fill("1A2A3A")

    row = 5
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = "Spécifications techniques"
    ws[f"A{row}"].font = Font(bold=True, color="FFFFFF", name="Calibri")
    ws[f"A{row}"].fill = _fill("34495E")
    row += 1
    for lbl, val in _flatten_specs(specs):
        ws[f"A{row}"] = lbl
        ws[f"B{row}"] = val
        ws[f"A{row}"].fill = _fill("F0F4F8")
        ws[f"A{row}"].font = Font(bold=True, name="Calibri", size=10)
        ws[f"B{row}"].font = Font(name="Calibri", size=10)
        for col in ("A", "B"):
            ws[f"{col}{row}"].border = _border()
        row += 1

    # Onglet Fournisseurs
    ws2 = wb.create_sheet("Fournisseurs & TCO")
    cols = ["Fournisseur", "Pays", "Email", "Prix unitaire (DZD)", "Délai", "TCO total (DZD)"]
    widths = [28, 14, 32, 22, 14, 22]
    for i, (col, w) in enumerate(zip(cols, widths), 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        c = ws2.cell(row=1, column=i, value=col)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill = _fill("34495E")
        c.border = _border()

    best = _best_supplier(suppliers)
    for ri, s in enumerate(suppliers, start=2):
        is_best = s.get("nom_fournisseur") == best
        vals = [s.get("nom_fournisseur", ""), PAYS, s.get("email", ""),  # PAYS STATIQUE
                float(s.get("prix_unitaire", 0)), _get_delai(s),
                float(s.get("tco_total", 0))]
        bg = "D4EDDA" if is_best else ("FFFFFF" if ri % 2 == 0 else "F0F4F8")
        for ci, v in enumerate(vals, 1):
            c = ws2.cell(row=ri, column=ci, value=v)
            c.font = Font(color="155724" if is_best else "2C3E50",
                          bold=is_best, name="Calibri", size=10)
            c.fill = _fill(bg)
            c.border = _border()
            if ci in (4, 6):
                c.number_format = "#,##0"

    # Onglet TCO
    ws3 = wb.create_sheet("TCO")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20
    kpis = [("TCO total (DZD)", finance.get("tco_total", 0)),
            ("Inflation moy. %", finance.get("inflation_pct", 0)),
            ("Maintenance / an", finance.get("maintenance_annuelle", 0)),
            ("ROI (3 ans) %", finance.get("roi_pct", 0)),
            ("VAN (DZD)", finance.get("van_dzd", 0)),
            ("TRI %", finance.get("tri_pct", 0))]
    for ri, (lbl, val) in enumerate(kpis, 1):
        ws3[f"A{ri}"] = lbl
        ws3[f"B{ri}"] = val
        ws3[f"A{ri}"].fill = _fill("F0F4F8")
        ws3[f"A{ri}"].font = Font(bold=True, name="Calibri", size=10)
        ws3[f"B{ri}"].number_format = "#,##0.00"
        for col in ("A", "B"):
            ws3[f"{col}{ri}"].border = _border()
    ws3.cell(row=len(kpis)+2, column=1, value="Année").font = Font(bold=True)
    ws3.cell(row=len(kpis)+2, column=2, value="Maintenance (DZD)").font = Font(bold=True)
    for r_i, bar in enumerate(_tco_bars(finance), start=len(kpis)+3):
        ws3.cell(row=r_i, column=1, value=bar["annee"])
        c = ws3.cell(row=r_i, column=2, value=bar["valeur"])
        c.number_format = "#,##0"
        c.fill = _fill("D1ECF1")

    # Onglet SWOT
    ws4 = wb.create_sheet("SWOT")
    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 60
    swot_rows = [
        ("Forces",       swot.get("forces", ""),       "EAF3DE", "3B6D11"),
        ("Faiblesses",   swot.get("faiblesses", ""),   "FAECE7", "993C1D"),
        ("Opportunités", swot.get("opportunites", ""), "E6F1FB", "185FA5"),
        ("Menaces",      swot.get("menaces", ""),      "FAEEDA", "854F0B"),
    ]
    for ri, (lbl, txt, bg, tc) in enumerate(swot_rows, 1):
        if isinstance(txt, list):
            txt = " · ".join(txt)
        ws4[f"A{ri}"] = lbl
        ws4[f"B{ri}"] = txt
        ws4[f"A{ri}"].font = Font(bold=True, color=tc, name="Calibri", size=10)
        ws4[f"A{ri}"].fill = _fill(bg)
        ws4[f"B{ri}"].fill = _fill(bg)
        ws4[f"B{ri}"].font = Font(name="Calibri", size=10)
        ws4[f"B{ri}"].alignment = Alignment(wrap_text=True)
        ws4.row_dimensions[ri].height = 40
        for col in ("A", "B"):
            ws4[f"{col}{ri}"].border = _border()

    wb.save(output_path)
    print(f"[M9] ✓ Excel : {output_path}")
    return output_path


# ── 4. JSON ───────────────────────────────────────────────────────────────
def generate_catalogue_json(specs, suppliers, finance, swot, output_path):
    best = _best_supplier(suppliers)
    data = {
        "generated_at": datetime.datetime.now().isoformat(),
        "pays": PAYS,  # STATIQUE
        "product": specs,
        "suppliers": [
            {
                "nom": s.get("nom_fournisseur"),
                "pays": PAYS,  # STATIQUE
                "email": s.get("email"),
                "prix_unitaire": s.get("prix_unitaire"),
                "delai": _get_delai(s),
                "tco_total": s.get("tco_total", 0),
                "meilleur": s.get("nom_fournisseur") == best,
            }
            for s in suppliers
        ],
        "finance": finance,
        "swot": {k: (v if isinstance(v, str) else " · ".join(v))
                 for k, v in swot.items()},
        "tco_bars": _tco_bars(finance),
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"[M9] ✓ JSON : {output_path}")
    return output_path


# ── 5. XML ────────────────────────────────────────────────────────────────
def generate_catalogue_xml(specs, suppliers, finance, swot, output_path):
    root = ET.Element("catalogue")
    ET.SubElement(root, "pays").text = PAYS  # STATIQUE
    ET.SubElement(root, "generated_at").text = datetime.datetime.now().isoformat()

    product = ET.SubElement(root, "product")
    for k, v in specs.items():
        if isinstance(v, dict):
            sub = ET.SubElement(product, k)
            for sk, sv in v.items():
                ET.SubElement(sub, sk).text = str(sv)
        else:
            ET.SubElement(product, k).text = str(v)

    best = _best_supplier(suppliers)
    suppliers_el = ET.SubElement(root, "suppliers")
    for s in suppliers[:6]:
        sup_el = ET.SubElement(suppliers_el, "supplier",
                               meilleur="true" if s.get("nom_fournisseur") == best else "false")
        ET.SubElement(sup_el, "name").text = s.get("nom_fournisseur", "")
        ET.SubElement(sup_el, "country").text = PAYS  # STATIQUE
        ET.SubElement(sup_el, "email").text = s.get("email", "")
        ET.SubElement(sup_el, "unit_price").text = str(s.get("prix_unitaire", 0))
        ET.SubElement(sup_el, "lead_time").text = _get_delai(s)
        ET.SubElement(sup_el, "tco_total").text = str(s.get("tco_total", 0))

    fin_el = ET.SubElement(root, "finance")
    for k, v in finance.items():
        ET.SubElement(fin_el, k).text = str(v)

    swot_el = ET.SubElement(root, "swot")
    for k, v in swot.items():
        ET.SubElement(swot_el, k).text = " · ".join(v) if isinstance(v, list) else str(v)

    xml_str = ET.tostring(root, encoding="utf-8")
    dom = minidom.parseString(xml_str)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(dom.toprettyxml(indent="  "))
    print(f"[M9] ✓ XML : {output_path}")
    return output_path


# ── 6. Point d'entrée LangGraph ──────────────────────────────────────────
def run_module_9(state: dict) -> dict:
    """
    Entrée : state['specs'], state['suppliers']/state['suppliers_negocie'],
             state['finance'], state['swot']
             Pays forcé à Algérie STATIQUE.
    Sortie : state['catalogue_files'] → liste des 5 fichiers générés
    """
    specs = state.get("specs", MOCK_SPECS)

    # Accepter suppliers OU suppliers_negocie OU all_tco
    raw_sup = (state.get("suppliers_negocie")
               or state.get("suppliers")
               or [])
    suppliers = _normalize_suppliers(raw_sup)

    # Récupérer TCO depuis all_tco si disponible
    all_tco = state.get("all_tco", [])
    tco_map = {
        item.get("supplier_info", {}).get("nom_fournisseur"): item.get("tco_result", {}).get("total_tco", 0)
        for item in all_tco
    }
    for s in suppliers:
        nom = s.get("nom_fournisseur")
        if nom and nom in tco_map and not s.get("tco_total"):
            s["tco_total"] = tco_map[nom]
        elif not s.get("tco_total"):
            s["tco_total"] = round(float(s.get("prix_unitaire", 5000)) * 200 * 2.2, 2)

    # Finance et SWOT
    finance = state.get("finance", MOCK_FINANCE)
    swot_raw = state.get("swot", MOCK_SWOT)
    # Normaliser SWOT : si valeur est liste → string
    swot = {}
    for k, v in swot_raw.items():
        swot[k] = " · ".join(v) if isinstance(v, list) else str(v)

    os.makedirs("outputs", exist_ok=True)

    pdf_path  = generate_catalogue_pdf(specs, suppliers, finance, swot,
                                        "outputs/catalogue.pdf")
    html_path = generate_catalogue_html(specs, suppliers, finance, swot,
                                         "outputs/catalogue.html")
    xlsx_path = generate_catalogue_excel(specs, suppliers, finance, swot,
                                          "outputs/catalogue.xlsx")
    json_path = generate_catalogue_json(specs, suppliers, finance, swot,
                                         "outputs/catalogue.json")
    xml_path  = generate_catalogue_xml(specs, suppliers, finance, swot,
                                        "outputs/catalogue.xml")

    state["catalogue_files"] = [pdf_path, html_path, xlsx_path, json_path, xml_path]
    return state


# ── Test indépendant ────────────────────────────────────────────────────
if __name__ == "__main__":
    def _load(path):
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        return None

    specs = _load("outputs/specs.json") or MOCK_SPECS
    neg = _load("outputs/negociation.json")
    suppliers = None
    if neg:
        suppliers = neg.get("suppliers_negocie") or [neg.get("best_supplier")]

    tco_data = _load("outputs/tco_data.json") or {}
    all_tco_raw = tco_data.get("suppliers", [])
    all_tco = [{"supplier_info": {"nom_fournisseur": s["nom"], "pays": PAYS},
                "tco_result": {"total_tco": s["total_tco"]}}
               for s in all_tco_raw]

    finance = _load("outputs/tco_data.json")
    if finance:
        finance = {
            "roi_pct": 34, "van_dzd": 280000, "tri_pct": 28,
            "tco_total": finance["suppliers"][0]["total_tco"] if finance.get("suppliers") else 847250,
            "inflation_pct": finance.get("inflation_moyenne_pct", 5.2),
            "maintenance_annuelle": 42000,
        }
    else:
        finance = MOCK_FINANCE

    # Correction : ne pas charger business_plan.pdf comme JSON
    # (ligne supprimée)

    result = run_module_9({
        "specs": specs,
        "suppliers": suppliers,
        "all_tco": all_tco,
        "finance": finance,
    })
    print("\n=== CATALOGUE GÉNÉRÉ ===")
    for f in result["catalogue_files"]:
        print(f"  ✓ {f}")