"""
Module 6 — Calcul TCO (Total Cost of Ownership) sur 10 ans
Utilise l'API World Bank pour l'inflation Algérie (DZ).
Pays = Algérie STATIQUE sur tous les fournisseurs.
Fallback mock complet si API indisponible.

Lancer seul :  python modules/module_6_tco.py
"""

import os
import json
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

# ── Constante pays ────────────────────────────────────────────────────────
PAYS = "Algérie"  # STATIQUE — ne jamais modifier

# ── Inflation mock Algérie (World Bank FP.CPI.TOTL.ZG, 2014-2023) ────────
INFLATION_MOCK_DZ = [2.9, 4.8, 6.4, 5.6, 4.3, 2.0, 3.0, 9.3, 9.3, 7.2]

# ── Fournisseurs mock ─────────────────────────────────────────────────────
MOCK_SUPPLIERS = [
    {
        "nom_fournisseur": "ValveTech SRL",
        "pays": PAYS,
        "email": "sales@valvetech.dz",
        "prix_unitaire": 4200,
        "delai_livraison": "18 jours",
    },
    {
        "nom_fournisseur": "HydroVal Algérie",
        "pays": PAYS,
        "email": "hydro@hydroval.dz",
        "prix_unitaire": 5500,
        "delai_livraison": "12 jours",
    },
    {
        "nom_fournisseur": "AlgéroFond SARL",
        "pays": PAYS,
        "email": "contact@algerofond.dz",
        "prix_unitaire": 4800,
        "delai_livraison": "14 jours",
    },
]


# ── 1. Récupération inflation World Bank ─────────────────────────────────
def get_inflation_algeria(years: int = 10) -> list:
    """
    Récupère le taux d'inflation annuel de l'Algérie via World Bank.
    Retourne INFLATION_MOCK_DZ si API indisponible.
    """
    if not HAS_REQUESTS:
        print("[M6] requests non installé → inflation mock")
        return INFLATION_MOCK_DZ[-years:]

    url = ("https://api.worldbank.org/v2/country/DZ/indicator/"
           "FP.CPI.TOTL.ZG?format=json&per_page=20")
    try:
        resp = requests.get(url, timeout=10)
        data = resp.json()
        records = data[1] if len(data) > 1 else []
        inflation = [float(r["value"]) for r in records if r.get("value") is not None]
        inflation.reverse()  # chronologique
        if len(inflation) >= years:
            return inflation[-years:]
        # Compléter si insuffisant
        last = inflation[-1] if inflation else 5.0
        inflation += [last] * (years - len(inflation))
        print(f"[M6] Inflation World Bank récupérée : {inflation}")
        return inflation
    except Exception as e:
        print(f"[M6] World Bank indisponible : {e} → inflation mock")
        return INFLATION_MOCK_DZ[-years:]


# ── 2. Calcul TCO ────────────────────────────────────────────────────────
def compute_tco_for_supplier(unit_price: float,
                              quantity: int = 200,
                              years: int = 10,
                              inflation_rates: list = None) -> dict:
    """
    Calcule le TCO complet sur `years` années.
    Structure :
      - Achat initial
      - Installation (15 % de l'achat)
      - Maintenance annuelle (8 % de l'achat × inflation cumulée)
      - Pièces de rechange (2 % de l'achat × inflation cumulée)
    """
    if inflation_rates is None:
        inflation_rates = [0.0] * years

    purchase = unit_price * quantity
    installation = round(purchase * 0.15, 2)
    cumul = purchase + installation

    maint_base = purchase * 0.08
    spare_base = purchase * 0.02

    breakdown = []
    cumul_inflation = 1.0

    for y in range(years):
        if y < len(inflation_rates):
            cumul_inflation *= (1 + inflation_rates[y] / 100)
        maint = round(maint_base * cumul_inflation, 2)
        spare = round(spare_base * cumul_inflation, 2)
        cumul = round(cumul + maint + spare, 2)
        breakdown.append({
            "year": 2025 + y,
            "maintenance": maint,
            "spare_parts": spare,
            "cumulative_tco": cumul,
            "inflation_rate": round(inflation_rates[y] if y < len(inflation_rates) else 0, 2)
        })

    return {
        "total_tco": cumul,
        "purchase": round(purchase, 2),
        "installation": round(installation, 2),
        "breakdown": breakdown
    }


# ── 3. Excel multi-onglets ───────────────────────────────────────────────
def _border():
    thin = Side(border_style="thin", color="BDC3C7")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _fill(hex_col: str):
    return PatternFill("solid", fgColor=hex_col)


def create_summary_sheet(wb, suppliers_results: list, quantity: int):
    """Onglet 1 : Tableau comparatif TCO."""
    ws = wb.create_sheet(title="Comparatif TCO", index=0)

    headers = ["Fournisseur", "Pays", "Email",
               "Prix unitaire (DZD)", "TCO total (DZD)",
               "Achat total (DZD)", "Installation (DZD)"]
    widths = [28, 14, 30, 22, 22, 22, 18]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill = _fill("34495E")
        c.alignment = Alignment(horizontal="center")
        c.border = _border()

    for ri, sup in enumerate(suppliers_results, start=2):
        info = sup["supplier_info"]
        tco = sup["tco_result"]
        vals = [
            info.get("nom_fournisseur", ""),
            PAYS,  # STATIQUE
            info.get("email", ""),
            info.get("prix_unitaire", 0),
            tco["total_tco"],
            tco["purchase"],
            tco["installation"],
        ]
        alt_bg = "F0F4F8" if ri % 2 == 0 else "FFFFFF"
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Calibri", size=10)
            c.fill = _fill(alt_bg)
            c.border = _border()
            if ci in (4, 5, 6, 7):
                c.number_format = "#,##0"
            if ci == 2:  # Pays en gras
                c.font = Font(bold=True, name="Calibri", size=10)


def create_supplier_sheet(wb, supplier_info: dict, tco_result: dict,
                           inflation_rates: list, quantity: int):
    """Onglet par fournisseur : détail TCO + graphique."""
    name = supplier_info.get("nom_fournisseur", "Fournisseur")
    sheet_name = name[:31]
    ws = wb.create_sheet(title=sheet_name)

    # En-tête
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Détail TCO — {name} ({PAYS})"  # STATIQUE
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    ws["A1"].fill = _fill("1A2A3A")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Colonnes données
    headers = ["Année", "Inflation (%)", "Maintenance (DZD)",
               "Pièces (DZD)", "TCO cumulé (DZD)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill = _fill("34495E")
        c.border = _border()
        ws.column_dimensions[get_column_letter(i)].width = 20

    for ri, yd in enumerate(tco_result["breakdown"], start=3):
        vals = [yd["year"], yd["inflation_rate"],
                yd["maintenance"], yd["spare_parts"], yd["cumulative_tco"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Calibri", size=10)
            c.fill = _fill("F0F4F8" if ri % 2 == 0 else "FFFFFF")
            c.border = _border()
            if ci >= 2:
                c.number_format = "#,##0.00"

    last_row = ws.max_row + 1
    ws.cell(row=last_row, column=1, value="TOTAL TCO")
    ws.cell(row=last_row, column=1).font = Font(bold=True, name="Calibri")
    c_total = ws.cell(row=last_row, column=5, value=tco_result["total_tco"])
    c_total.font = Font(bold=True, name="Calibri")
    c_total.number_format = "#,##0"

    # Graphique barres TCO cumulé
    chart = BarChart()
    chart.title = f"TCO cumulé — {name}"
    chart.x_axis.title = "Année"
    chart.y_axis.title = "DZD"
    data = Reference(ws, min_col=5, max_col=5, min_row=2,
                     max_row=2 + len(tco_result["breakdown"]))
    categories = Reference(ws, min_col=1, min_row=3,
                           max_row=2 + len(tco_result["breakdown"]))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, "G3")

    # Onglet paramètres
    ws_p = wb.create_sheet(title=f"Params_{sheet_name[:18]}")
    params = [
        ["Paramètre", "Valeur"],
        ["Fournisseur", name],
        ["Pays", PAYS],  # STATIQUE
        ["Email", supplier_info.get("email", "")],
        ["Prix unitaire (DZD)", supplier_info.get("prix_unitaire", 0)],
        ["Quantité", quantity],
        ["Achat total (DZD)", tco_result["purchase"]],
        ["Installation 15% (DZD)", tco_result["installation"]],
        ["Inflation moy. (%)", round(sum(inflation_rates) / len(inflation_rates), 2)
         if inflation_rates else 0],
    ]
    for ri, row in enumerate(params, 1):
        for ci, v in enumerate(row, 1):
            c = ws_p.cell(row=ri, column=ci, value=v)
            if ri == 1:
                c.font = Font(bold=True, name="Calibri")
            c.border = _border()
        ws_p.column_dimensions["A"].width = 30
        ws_p.column_dimensions["B"].width = 25


def generate_tco_excel(suppliers_results: list, inflation_rates: list,
                        quantity: int, output_path: str) -> str:
    """Génère le fichier Excel TCO multi-onglets."""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    create_summary_sheet(wb, suppliers_results, quantity)
    for res in suppliers_results:
        create_supplier_sheet(wb, res["supplier_info"], res["tco_result"],
                              inflation_rates, quantity)
    wb.save(output_path)
    print(f"[M6] ✓ Excel TCO : {output_path}")
    return output_path


# ── 4. JSON global ───────────────────────────────────────────────────────
def generate_tco_json(suppliers_results: list, inflation_rates: list,
                      quantity: int, output_path: str) -> str:
    data = {
        "pays": PAYS,  # STATIQUE
        "quantity": quantity,
        "inflation_rates": inflation_rates,
        "inflation_moyenne_pct": round(sum(inflation_rates) / len(inflation_rates), 2)
        if inflation_rates else 0,
        "suppliers": [
            {
                "nom": s["supplier_info"].get("nom_fournisseur"),
                "pays": PAYS,  # STATIQUE
                "email": s["supplier_info"].get("email", ""),
                "prix_unitaire": s["supplier_info"].get("prix_unitaire", 0),
                "total_tco": s["tco_result"]["total_tco"],
                "purchase": s["tco_result"]["purchase"],
                "installation": s["tco_result"]["installation"],
                "breakdown": s["tco_result"]["breakdown"],
            }
            for s in suppliers_results
        ],
    }
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"[M6] ✓ JSON TCO : {output_path}")
    return output_path


# ── 5. Point d'entrée LangGraph ──────────────────────────────────────────
def run_module_6(state: dict) -> dict:
    """
    Entrée  : state['suppliers'] → liste de fournisseurs (depuis M4/M5)
              Pays forcé à Algérie sur tous.
    Sortie  : state['tco'], state['all_tco'],
              state['tco_excel_path'], state['tco_json_path']
    """
    suppliers = state.get("suppliers", [])
    if not suppliers:
        print("[M6] Aucun fournisseur → données mock")
        suppliers = MOCK_SUPPLIERS

    # Forcer pays = Algérie sur tous les fournisseurs
    for s in suppliers:
        s["pays"] = PAYS

    quantity = state.get("quantite", 200)
    years = state.get("years", 10)

    # Récupérer inflation
    inflation_rates = get_inflation_algeria(years)

    # Calculer TCO pour chaque fournisseur
    suppliers_results = []
    for sup in suppliers:
        unit_price = float(sup.get("prix_unitaire", 12500))
        tco_result = compute_tco_for_supplier(unit_price, quantity, years, inflation_rates)
        suppliers_results.append({
            "supplier_info": {**sup, "pays": PAYS},  # STATIQUE
            "tco_result": tco_result
        })
        print(f"[M6] TCO {sup.get('nom_fournisseur', '?')} : {tco_result['total_tco']:,.0f} DZD")

    # Générer les fichiers
    os.makedirs("outputs", exist_ok=True)
    excel_path = "outputs/tco_report.xlsx"
    json_path = "outputs/tco_data.json"

    generate_tco_excel(suppliers_results, inflation_rates, quantity, excel_path)
    generate_tco_json(suppliers_results, inflation_rates, quantity, json_path)

    # Mettre à jour le state
    state["tco"] = suppliers_results[0]["tco_result"] if suppliers_results else {}
    state["all_tco"] = suppliers_results
    state["tco_excel_path"] = excel_path
    state["tco_json_path"] = json_path

    # Calcul inflation moyenne pour M7
    state["inflation_moyenne"] = (
        round(sum(inflation_rates) / len(inflation_rates), 2) if inflation_rates else 5.2
    )

    return state


# ── Test indépendant ────────────────────────────────────────────────────
if __name__ == "__main__":
    # Charger depuis M5 si disponible
    neg_path = "outputs/negociation.json"
    m4_path = "outputs/fournisseurs.json"

    suppliers = None
    if os.path.exists(neg_path):
        with open(neg_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        suppliers = data.get("suppliers_negocie") or data.get("best_supplier")
        if isinstance(suppliers, dict):
            suppliers = [suppliers]
        print(f"[M6] {len(suppliers)} fournisseurs chargés depuis M5")
    elif os.path.exists(m4_path):
        with open(m4_path, "r", encoding="utf-8") as f:
            suppliers = json.load(f)
        print(f"[M6] {len(suppliers)} fournisseurs chargés depuis M4")

    result = run_module_6({"suppliers": suppliers} if suppliers else {})
    print(f"\nExcel : {result['tco_excel_path']}")
    print(f"JSON  : {result['tco_json_path']}")
    print(f"TCO total (premier fournisseur) : {result['tco']['total_tco']:,.0f} DZD")