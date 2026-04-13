"""
Module 7 — Business Plan Professionnel (PDF + Excel)
Pays = Algérie STATIQUE.
Accepte : state['suppliers'] liste OU un seul supplier (dict).
SWOT généré par Mistral (Ollama) avec fallback mock robuste.
Correction : ajout du '#' devant les codes de couleur manquants.
"""

import os
import json
import re
import datetime

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                 Table, TableStyle, PageBreak, HRFlowable)

try:
    import ollama
    HAS_OLLAMA = True
except ImportError:
    HAS_OLLAMA = False

# ── Constante pays ─────────────────────────────────────────────────────────
PAYS = "Algérie"  # STATIQUE

# ── Données mock ────────────────────────────────────────────────────────────
MOCK_SPECS = {
    "diametre_nominal": "DN100",
    "pression_nominale": "PN40",
    "materiau": "Inox 316L"
}

MOCK_SUPPLIERS = [
    {"nom_fournisseur": "ValveTech SRL", "pays": PAYS, "email": "sales@valvetech.dz",
     "prix_unitaire": 4200, "delai_livraison": "18 jours"},
    {"nom_fournisseur": "HydroVal Algérie", "pays": PAYS, "email": "hydro@hydroval.dz",
     "prix_unitaire": 5500, "delai_livraison": "12 jours"},
]

MOCK_SWOT = {
    "forces": [
        "Matériau premium Inox 316L certifié EN 558",
        "Réseau fournisseurs algériens établi",
        "Délais de livraison compétitifs (10-18 jours)"
    ],
    "faiblesses": [
        "Dépendance partielle aux matières premières importées",
        "Coût unitaire élevé vs concurrence asiatique",
        "Capacité de production initiale limitée"
    ],
    "opportunites": [
        "Croissance du secteur pétro-gaz en Algérie (+8% prévu)",
        "Programme de substitution aux importations (MDPI)",
        "Demande croissante en équipements haute pression"
    ],
    "menaces": [
        "Fluctuation du dinar algérien (DZD)",
        "Pression des importateurs asiatiques à bas coût",
        "Complexité des procédures douanières"
    ]
}


# ── 1. Calculs financiers ────────────────────────────────────────────────
def compute_financials(initial_investment: float,
                        cashflows: list,
                        discount_rate: float = 0.12) -> dict:
    """Calcule NPV, ROI, IRR."""
    npv = sum(cf / ((1 + discount_rate) ** (i + 1))
              for i, cf in enumerate(cashflows)) - initial_investment
    total_gain = sum(cashflows)
    roi = ((total_gain - initial_investment) / initial_investment * 100
           if initial_investment else 0)

    # TRI approximé par Newton-Raphson (3 itérations)
    irr = discount_rate
    for _ in range(50):
        npv_iter = sum(cf / ((1 + irr) ** (i + 1))
                       for i, cf in enumerate(cashflows)) - initial_investment
        dnpv = sum(-(i + 1) * cf / ((1 + irr) ** (i + 2))
                   for i, cf in enumerate(cashflows))
        if abs(dnpv) < 1e-10:
            break
        irr -= npv_iter / dnpv
    irr = max(0, min(irr, 1))  # borner entre 0 et 100%

    return {
        "npv": round(npv, 2),
        "roi": round(roi, 2),
        "irr": round(irr * 100, 1),
        "investment": round(initial_investment, 2)
    }


def generate_projections(unit_price: float,
                          quantities: list = None,
                          tco_breakdown_first3: list = None) -> list:
    """Projections 3 ans : CA, charges, bénéfice."""
    if quantities is None:
        quantities = [200, 500, 1000]
    selling_price = unit_price * 1.5
    years = [2025, 2026, 2027]
    cumulative = 0
    projections = []
    for i, qty in enumerate(quantities):
        revenue = qty * selling_price
        if tco_breakdown_first3 and i < len(tco_breakdown_first3):
            op_costs = (tco_breakdown_first3[i].get("maintenance", 0) +
                        tco_breakdown_first3[i].get("spare_parts", 0))
        else:
            op_costs = unit_price * 200 * 0.10
        profit = revenue - op_costs
        cumulative += profit
        projections.append({
            "year": years[i] if i < len(years) else 2025 + i,
            "units_sold": qty,
            "revenue": round(revenue, 2),
            "operating_costs": round(op_costs, 2),
            "profit": round(profit, 2),
            "cumulative_profit": round(cumulative, 2)
        })
    return projections


# ── 2. SWOT (Ollama Mistral avec fallback robuste) ───────────────────────
def generate_swot(product_description: str) -> dict:
    """
    Génère un SWOT via Mistral. Retourne MOCK_SWOT si erreur.
    """
    if not HAS_OLLAMA:
        print("[M7] Ollama non disponible → SWOT mock")
        return MOCK_SWOT

    prompt = f"""Tu es un expert en stratégie industrielle algérienne.
Pour le produit : {product_description}
Génère une analyse SWOT en JSON strictement ainsi :
{{
    "forces": ["force1", "force2", "force3"],
    "faiblesses": ["faiblesse1", "faiblesse2", "faiblesse3"],
    "opportunites": ["opportunite1", "opportunite2", "opportunite3"],
    "menaces": ["menace1", "menace2", "menace3"]
}}
Contexte : marché algérien, industrie pétro-gaz, DZD.
Réponds UNIQUEMENT avec le JSON, sans backticks.
"""
    try:
        response = ollama.chat(
            model="mistral",
            messages=[{"role": "user", "content": prompt}]
        )
        content = response["message"]["content"].strip()
        content = re.sub(r"```json\s*", "", content)
        content = re.sub(r"```\s*", "", content)
        start = content.find("{")
        end = content.rfind("}") + 1
        if start < 0 or end <= start:
            raise ValueError("JSON introuvable")
        swot = json.loads(content[start:end])

        # Normaliser les clés anglaises → françaises
        key_map = {
            "strengths": "forces", "weaknesses": "faiblesses",
            "opportunities": "opportunites", "threats": "menaces"
        }
        normalized = {}
        for k, v in swot.items():
            k_lower = k.lower().strip()
            normalized[key_map.get(k_lower, k_lower)] = v

        required = ["forces", "faiblesses", "opportunites", "menaces"]
        if all(k in normalized for k in required):
            return normalized
        raise ValueError(f"Clés manquantes : {set(required) - set(normalized)}")

    except Exception as e:
        print(f"[M7] Erreur SWOT Ollama : {e} → mock")
        return MOCK_SWOT


# ── 3. Excel ─────────────────────────────────────────────────────────────
def _border():
    t = Side(border_style="thin", color="BDC3C7")
    return Border(left=t, right=t, top=t, bottom=t)


def _fill(h):
    return PatternFill("solid", fgColor=h)


def generate_business_plan_excel(suppliers_data: list, output_path: str) -> str:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Feuille comparaison
    ws_comp = wb.create_sheet("Comparaison fournisseurs")
    comp_headers = ["Fournisseur", "Pays", "Email", "Prix unit. (DZD)",
                    "TCO total (DZD)", "VAN (DZD)", "ROI (%)", "TRI (%)"]
    widths_comp = [28, 14, 30, 20, 22, 22, 12, 12]
    for i, (h, w) in enumerate(zip(comp_headers, widths_comp), 1):
        ws_comp.column_dimensions[get_column_letter(i)].width = w
        c = ws_comp.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill = _fill("34495E")
        c.border = _border()

    best_roi = max((s["financials"]["roi"] for s in suppliers_data), default=0)
    for ri, sup in enumerate(suppliers_data, start=2):
        is_best = sup["financials"]["roi"] == best_roi
        vals = [
            sup["name"],
            PAYS,  # STATIQUE
            sup.get("email", ""),
            sup["unit_price"],
            sup["tco_total"],
            sup["financials"]["npv"],
            sup["financials"]["roi"],
            sup["financials"]["irr"],
        ]
        bg = "D4EDDA" if is_best else ("F0F4F8" if ri % 2 == 0 else "FFFFFF")
        for ci, v in enumerate(vals, 1):
            c = ws_comp.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Calibri", size=10,
                          bold=is_best, color="155724" if is_best else "2C3E50")
            c.fill = _fill(bg)
            c.border = _border()
            if ci in (4, 5, 6):
                c.number_format = "#,##0"
            if ci in (7, 8):
                c.number_format = "0.0"

    # Feuilles individuelles
    for sup in suppliers_data:
        sheet_name = sup["name"][:31]
        ws = wb.create_sheet(sheet_name)
        proj_headers = ["Année", "Unités vendues", "CA (DZD)",
                        "Charges (DZD)", "Bénéfice (DZD)", "Bénéf. cumulé (DZD)"]
        for i, h in enumerate(proj_headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            c.fill = _fill("34495E")
            c.border = _border()
            ws.column_dimensions[get_column_letter(i)].width = 20

        for ri, p in enumerate(sup["projections"], start=2):
            vals = [p["year"], p["units_sold"], p["revenue"],
                    p["operating_costs"], p["profit"], p["cumulative_profit"]]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = _border()
                if ci >= 3:
                    c.number_format = "#,##0"

        # Graphique
        chart = LineChart()
        chart.title = f"CA & Bénéfice — {sup['name']}"
        chart.x_axis.title = "Année"
        data = Reference(ws, min_col=3, max_col=5, min_row=1,
                         max_row=len(sup["projections"]) + 1)
        categories = Reference(ws, min_col=1, min_row=2,
                                max_row=len(sup["projections"]) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        ws.add_chart(chart, "H5")

    wb.save(output_path)
    print(f"[M7] ✓ Excel Business Plan : {output_path}")
    return output_path


# ── 4. PDF ReportLab ──────────────────────────────────────────────────────
def _header_footer(canvas_obj, doc):
    canvas_obj.saveState()
    canvas_obj.setFont("Helvetica-Bold", 8)
    canvas_obj.setFillColor(colors.HexColor("#2c3e50"))
    canvas_obj.drawString(20 * mm, A4[1] - 12 * mm,
                          f"INDUSTRIE IA — Business Plan — {PAYS}")
    canvas_obj.drawCentredString(A4[0] / 2, 12 * mm, f"Page {doc.page}")
    canvas_obj.restoreState()


def generate_business_plan_pdf(specs: dict, suppliers_data: list,
                                output_path: str) -> str:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        rightMargin=20 * mm, leftMargin=20 * mm,
        topMargin=25 * mm, bottomMargin=25 * mm
    )
    styles = getSampleStyleSheet()
    W = 170 * mm

    title_s = ParagraphStyle("Title", parent=styles["Normal"],
                              fontSize=20, textColor=colors.HexColor("#2c3e50"),
                              fontName="Helvetica-Bold", alignment=1, spaceAfter=12)
    h2 = ParagraphStyle("H2", parent=styles["Normal"],
                         fontSize=13, textColor=colors.HexColor("#2980b9"),
                         fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=6)
    h3 = ParagraphStyle("H3", parent=styles["Normal"],
                         fontSize=11, textColor=colors.HexColor("#34495e"),
                         fontName="Helvetica-Bold", spaceBefore=8, spaceAfter=4)
    normal = ParagraphStyle("N", parent=styles["Normal"],
                             fontSize=9, spaceAfter=3)
    footer_s = ParagraphStyle("Foot", parent=styles["Normal"],
                               fontSize=8, textColor=colors.HexColor("#7f8c8d"),
                               alignment=1)

    story = []
    dn = specs.get("diametre_nominal", "DN100")
    pn = specs.get("pression_nominale", "PN40")
    mat = specs.get("materiau", "Inox 316L")
    if isinstance(mat, dict):
        mat = mat.get("disque", "Inox 316L")

    # ── Page de couverture ──────────────────────────────────────────────
    story.append(Spacer(1, 40 * mm))
    story.append(Paragraph(
        f"Business Plan — Vanne {dn} {pn} ({mat})", title_s))
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        f"Pays : <b>{PAYS}</b> · Date : {datetime.datetime.now().strftime('%d/%m/%Y')}",
        ParagraphStyle("sub", parent=styles["Normal"], fontSize=11,
                       alignment=1, textColor=colors.HexColor("#7f8c8d"))))
    story.append(Spacer(1, 8 * mm))

    # Mini tableau de synthèse couverture
    cov_data = [
        ["Produit", f"Vanne {dn} {pn}"],
        ["Matériau", mat],
        ["Pays d'opération", PAYS],
        ["Fournisseurs analysés", str(len(suppliers_data))],
        ["Période de projection", "3 ans (2025–2027)"],
        ["Taux d'actualisation", "12 %"],
    ]
    cov_table = Table(cov_data, colWidths=[70 * mm, W - 70 * mm])
    cov_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f4f8")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#bdc3c7")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(cov_table)
    story.append(PageBreak())

    # ── Tableau comparatif ──────────────────────────────────────────────
    story.append(Paragraph("1. Comparaison des fournisseurs", h2))
    comp_headers = ["Fournisseur", "Pays", "Prix unit.\n(DZD)",
                    "TCO total\n(DZD)", "VAN (DZD)", "ROI (%)"]
    comp_data = [[Paragraph(f"<b>{h}</b>", normal) for h in comp_headers]]
    for sup in suppliers_data:
        comp_data.append([
            Paragraph(f"<b>{sup['name']}</b>", normal),
            Paragraph(PAYS, normal),
            Paragraph(f"{sup['unit_price']:,.0f}", normal),
            Paragraph(f"{sup['tco_total']:,.0f}", normal),
            Paragraph(f"{sup['financials']['npv']:,.0f}", normal),
            Paragraph(f"{sup['financials']['roi']:.1f} %", normal),
        ])
    col_widths = [45 * mm, 22 * mm, 25 * mm, 28 * mm, 25 * mm, 18 * mm]
    comp_table = Table(comp_data, colWidths=col_widths, repeatRows=1)
    comp_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bdc3c7")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f8f9fa")]),
    ]))
    story.append(comp_table)
    story.append(Spacer(1, 8 * mm))

    # ── Sections par fournisseur ────────────────────────────────────────
    for sup in suppliers_data:
        story.append(PageBreak())
        story.append(Paragraph(
            f"2. Analyse détaillée — {sup['name']} ({PAYS})", h2))
        story.append(Paragraph(
            f"Prix unitaire : <b>{sup['unit_price']:,.0f} DZD</b> · "
            f"TCO total : <b>{sup['tco_total']:,.0f} DZD</b>",
            normal))
        story.append(Spacer(1, 5 * mm))

        # SWOT
        story.append(Paragraph("Analyse SWOT", h3))
        swot = sup.get("swot", MOCK_SWOT)
        swot_configs = [
            ("Forces",       swot.get("forces", []),       "#eaf3de", "#3b6d11"),
            ("Faiblesses",   swot.get("faiblesses", []),   "#faece7", "#993c1d"),
            ("Opportunités", swot.get("opportunites", []), "#e6f1fb", "#185fa5"),
            ("Menaces",      swot.get("menaces", []),      "#faeeda", "#854f0b"),
        ]
        half = W / 2 - 2 * mm
        swot_rows = []
        for i in range(0, 4, 2):
            row = []
            for j in range(2):
                lbl, items, bg, tc = swot_configs[i + j]
                content = [Paragraph(f"<b>{lbl}</b>",
                                     ParagraphStyle("sl", parent=normal,
                                                    textColor=colors.HexColor(tc)))]
                if isinstance(items, list):
                    for item in items:
                        content.append(Paragraph(f"• {item}", normal))
                else:
                    content.append(Paragraph(str(items), normal))

                cell = Table([[p] for p in content], colWidths=[half])
                cell.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(bg)),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ]))
                row.append(cell)
            swot_rows.append(row)
        swot_table = Table(swot_rows, colWidths=[half, half])
        swot_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(swot_table)
        story.append(Spacer(1, 6 * mm))

        # Projections financières
        story.append(Paragraph("Projections financières (3 ans)", h3))
        proj_headers = ["Année", "Unités", "CA (DZD)",
                        "Charges (DZD)", "Bénéfice (DZD)", "Bénéf. cumulé (DZD)"]
        proj_data = [[Paragraph(f"<b>{h}</b>", normal) for h in proj_headers]]
        for p in sup["projections"]:
            proj_data.append([
                Paragraph(str(p["year"]), normal),
                Paragraph(str(p["units_sold"]), normal),
                Paragraph(f"{p['revenue']:,.0f}", normal),
                Paragraph(f"{p['operating_costs']:,.0f}", normal),
                Paragraph(f"{p['profit']:,.0f}", normal),
                Paragraph(f"{p['cumulative_profit']:,.0f}", normal),
            ])
        proj_table = Table(proj_data, colWidths=[18 * mm, 18 * mm, 30 * mm,
                                                  28 * mm, 30 * mm, 30 * mm],
                           repeatRows=1)
        proj_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bdc3c7")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f8f9fa")]),
        ]))
        story.append(proj_table)
        story.append(Spacer(1, 5 * mm))

        # KPIs financiers
        fin = sup["financials"]
        kpi_data = [
            [Paragraph(f"<b>{fin['roi']:.1f} %</b>", normal),
             Paragraph(f"<b>{fin['npv']:+,.0f} DZD</b>", normal),
             Paragraph(f"<b>{fin['irr']:.1f} %</b>", normal)],
            [Paragraph("ROI (3 ans)", normal),
             Paragraph("VAN (12 %)", normal),
             Paragraph("TRI", normal)],
        ]
        kpi_table = Table(kpi_data, colWidths=[W / 3] * 3)
        kpi_style = [
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d0d8e0")),
        ]
        # CORRECTION : ajout du '#' devant les codes couleur
        kpi_colors = ["#D4EDDA", "#D1ECF1", "#F0F4F8"]
        for col, bg in enumerate(kpi_colors):
            kpi_style.append(("BACKGROUND", (col, 0), (col, -1),
                               colors.HexColor(bg)))
        kpi_table.setStyle(TableStyle(kpi_style))
        story.append(kpi_table)

    # ── Pied de page final ──────────────────────────────────────────────
    story.append(Spacer(1, 8 * mm))
    story.append(HRFlowable(width=W, color=colors.HexColor("#bdc3c7"), thickness=0.5))
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph(
        f"Document généré automatiquement — INDUSTRIE IA · OpenIndustry {PAYS} · "
        f"{datetime.datetime.now().strftime('%d/%m/%Y')}",
        footer_s
    ))

    doc.build(story,
              onFirstPage=_header_footer,
              onLaterPages=_header_footer)
    print(f"[M7] ✓ PDF Business Plan : {output_path}")
    return output_path


# ── 5. Point d'entrée LangGraph ──────────────────────────────────────────
def run_module_7(state: dict) -> dict:
    """
    Entrée  : state['specs'], state['suppliers'] (liste OU dict unique),
              state['tco'] (optionnel)
    Sortie  : state['business_plan_pdf'], state['business_plan_excel'],
              state['suppliers_business_plans']
    Pays = Algérie STATIQUE sur tous les fournisseurs.
    """
    specs = state.get("specs", MOCK_SPECS)

    # ── Normaliser suppliers : accepte liste OU dict unique ─────────────
    raw_suppliers = state.get("suppliers", [])
    if isinstance(raw_suppliers, dict):
        raw_suppliers = [raw_suppliers]  # Un seul → liste de 1
    if not raw_suppliers:
        raw_suppliers = MOCK_SUPPLIERS

    # Forcer pays = Algérie
    for s in raw_suppliers:
        s["pays"] = PAYS

    tco_data = state.get("tco", {})
    tco_breakdown = tco_data.get("breakdown", []) if isinstance(tco_data, dict) else []

    # ── Récupérer TCO par fournisseur depuis all_tco ──────────────────
    all_tco = state.get("all_tco", [])
    tco_map = {}
    for item in all_tco:
        nom = item.get("supplier_info", {}).get("nom_fournisseur")
        if nom:
            tco_map[nom] = item.get("tco_result", {}).get("total_tco", 0)

    mat = specs.get("materiau", "Inox 316L")
    if isinstance(mat, dict):
        mat = mat.get("disque", "Inox 316L")

    suppliers_data = []
    for sup in raw_suppliers:
        unit_price = float(sup.get("prix_unitaire", 12500))
        nom = sup.get("nom_fournisseur", "Inconnu")

        product_desc = (
            f"vanne {specs.get('diametre_nominal', 'DN100')} "
            f"{specs.get('pression_nominale', 'PN40')} en {mat} "
            f"— fournisseur {nom} — marché {PAYS}"
        )
        swot = generate_swot(product_desc)

        projections = generate_projections(unit_price, tco_breakdown_first3=tco_breakdown[:3])
        initial_investment = unit_price * 200 * 1.15
        cashflows = [p["profit"] for p in projections]
        financials = compute_financials(initial_investment, cashflows)

        tco_total = tco_map.get(nom) or tco_data.get("total_tco") or round(unit_price * 200 * 2.2, 2)

        suppliers_data.append({
            "name": nom,
            "pays": PAYS,  # STATIQUE
            "email": sup.get("email", ""),
            "unit_price": unit_price,
            "tco_total": tco_total,
            "swot": swot,
            "projections": projections,
            "financials": financials
        })

    os.makedirs("outputs", exist_ok=True)
    excel_path = "outputs/business_plan_projections.xlsx"
    pdf_path = "outputs/business_plan.pdf"

    generate_business_plan_excel(suppliers_data, excel_path)
    generate_business_plan_pdf(specs, suppliers_data, pdf_path)

    state["business_plan_pdf"] = pdf_path
    state["business_plan_excel"] = excel_path
    state["suppliers_business_plans"] = suppliers_data

    # Extraire les indicateurs financiers pour M9
    if suppliers_data:
        best = max(suppliers_data, key=lambda x: x["financials"]["roi"])
        state["finance"] = {
            "roi_pct": best["financials"]["roi"],
            "van_dzd": best["financials"]["npv"],
            "tri_pct": best["financials"]["irr"],
            "tco_total": best["tco_total"],
            "inflation_pct": state.get("inflation_moyenne", 5.2),
            "maintenance_annuelle": round(best["unit_price"] * 200 * 0.08, 0),
        }

    return state


# ── Test indépendant ────────────────────────────────────────────────────
if __name__ == "__main__":
    # Charger depuis modules précédents si disponibles
    def _load(path):
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        return None

    specs = _load("outputs/specs.json") or MOCK_SPECS
    neg = _load("outputs/negociation.json")
    suppliers = None
    if neg:
        suppliers = neg.get("suppliers_negocie") or [neg.get("best_supplier")]

    tco_data = _load("outputs/tco_data.json") or {}
    all_tco_raw = tco_data.get("suppliers", [])
    all_tco = [{"supplier_info": {"nom_fournisseur": s["nom"], "pays": PAYS,
                                   "email": s.get("email", ""),
                                   "prix_unitaire": s["prix_unitaire"]},
                "tco_result": {"total_tco": s["total_tco"],
                               "breakdown": s.get("breakdown", [])}}
               for s in all_tco_raw]

    result = run_module_7({
        "specs": specs,
        "suppliers": suppliers,
        "all_tco": all_tco,
        "tco": {"total_tco": all_tco[0]["tco_result"]["total_tco"],
                "breakdown": all_tco[0]["tco_result"]["breakdown"]} if all_tco else {}
    })

    print(f"\nPDF : {result['business_plan_pdf']}")
    print(f"Excel : {result['business_plan_excel']}")